from __future__ import annotations

import argparse
import csv
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]

REQUIRED_COLUMNS = (
    "row_id",
    "ref",
    "hebrew_phrase",
    "linear_translation",
    "metsudah_context",
    "koren_context",
    "skill_primary",
    "skill_secondary",
    "current_status",
    "issue_type",
    "what_to_check",
    "recommended_default_decision",
    "yossi_decision",
    "yossi_notes",
)

ALLOWED_DECISIONS = (
    "verified",
    "fix_translation",
    "fix_hebrew_phrase",
    "fix_phrase_boundary",
    "fix_skill_classification",
    "source_only",
    "block_for_questions",
    "needs_follow_up",
)

ISSUE_SORT_ORDER = {
    "missing_translation": 0,
    "needs_follow_up": 1,
    "long_parenthetical": 2,
    "long_hebrew_boundary": 3,
    "awkward_source_wording": 4,
    "phrase_boundary_check": 5,
    "classification_check": 6,
    "source_only_recommended": 7,
    "clean_sample": 8,
}

LONG_HEBREW_THRESHOLD = 45
LONG_TRANSLATION_THRESHOLD = 80
AWKWARD_MARKERS = ("spread", "small(er)", "big luminary", "made for", "face of", "was supposed")


def repo_relative(path: Path) -> str:
    try:
        return path.relative_to(ROOT).as_posix()
    except ValueError:
        return path.as_posix()


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def has_parenthetical(row: dict[str, str]) -> bool:
    return any("(" in row.get(field, "") or ")" in row.get(field, "") for field in translation_fields())


def translation_fields() -> tuple[str, str, str]:
    return ("alternate_translation", "source_translation_metsudah", "secondary_translation_koren")


def missing_translation(row: dict[str, str]) -> bool:
    return any(not row.get(field, "").strip() for field in translation_fields())


def long_parenthetical(row: dict[str, str]) -> bool:
    return has_parenthetical(row) and any(len(row.get(field, "")) >= LONG_TRANSLATION_THRESHOLD for field in translation_fields())


def long_hebrew(row: dict[str, str]) -> bool:
    return len(row.get("hebrew_word_or_phrase", "")) >= LONG_HEBREW_THRESHOLD


def awkward_wording(row: dict[str, str]) -> bool:
    text = row.get("alternate_translation", "").lower()
    return any(marker in text for marker in AWKWARD_MARKERS)


def phrase_boundary_check(row: dict[str, str]) -> bool:
    uncertainty = row.get("uncertainty_reason", "").lower()
    return "phrase alignment" in uncertainty or "phrase" in uncertainty and "review" in uncertainty


def classification_check(row: dict[str, str]) -> bool:
    return not row.get("skill_primary", "").strip() or not row.get("skill_secondary", "").strip()


def issue_types_for_row(row: dict[str, str]) -> list[str]:
    issues: list[str] = []
    if missing_translation(row):
        issues.append("missing_translation")
    if long_parenthetical(row):
        issues.append("long_parenthetical")
    if long_hebrew(row):
        issues.append("long_hebrew_boundary")
    if awkward_wording(row):
        issues.append("awkward_source_wording")
    if phrase_boundary_check(row):
        issues.append("phrase_boundary_check")
    if classification_check(row):
        issues.append("classification_check")
    if row.get("question_allowed") == "needs_review":
        issues.append("source_only_recommended")
    if not issues:
        issues.append("clean_sample")
    return sorted(dict.fromkeys(issues), key=lambda issue: ISSUE_SORT_ORDER[issue])


def what_to_check_for_issues(issues: list[str]) -> str:
    checks: list[str] = []
    if "missing_translation" in issues:
        checks.append("Confirm whether missing source/translation context blocks extraction verification.")
    if "long_parenthetical" in issues:
        checks.append("Confirm the parenthetical belongs to this Hebrew phrase.")
    if "long_hebrew_boundary" in issues:
        checks.append("Confirm the Hebrew phrase boundary and segment join are reasonable.")
    if "awkward_source_wording" in issues:
        checks.append("Confirm awkward English is source-derived and should remain unchanged.")
    if "phrase_boundary_check" in issues:
        checks.append("Confirm Hebrew-English phrase alignment against the trusted source.")
    if "classification_check" in issues:
        checks.append("Confirm skill/classification is reasonable for planning.")
    if "source_only_recommended" in issues:
        checks.append("Keep source-only unless a separate future question/protected-preview gate approves use.")
    if "clean_sample" in issues:
        checks.append("Representative clean sample; confirm extraction alignment.")
    return " ".join(checks)


def recommended_decision(issues: list[str]) -> str:
    if "missing_translation" in issues or "needs_follow_up" in issues:
        return "needs_follow_up"
    if issues == ["source_only_recommended"]:
        return "source_only"
    return "verified"


def sort_key(item: dict[str, str]) -> tuple[int, str, str]:
    first_issue = item["issue_type"].split(";")[0]
    return (ISSUE_SORT_ORDER.get(first_issue, 99), item["ref"], item["row_id"])


def build_review_items(rows: list[dict[str, str]], map_path: Path, max_clean_samples: int) -> list[dict[str, str]]:
    items: list[dict[str, str]] = []
    clean_seen = 0
    stem = map_path.stem
    for index, row in enumerate(rows, start=1):
        issues = issue_types_for_row(row)
        if issues == ["clean_sample"]:
            clean_seen += 1
            if clean_seen > max_clean_samples:
                continue
        items.append(
            {
                "row_id": f"{stem}_row_{index:03d}",
                "ref": row.get("ref", ""),
                "hebrew_phrase": row.get("hebrew_word_or_phrase", ""),
                "linear_translation": row.get("alternate_translation", ""),
                "metsudah_context": row.get("source_translation_metsudah", ""),
                "koren_context": row.get("secondary_translation_koren", ""),
                "skill_primary": row.get("skill_primary", ""),
                "skill_secondary": row.get("skill_secondary", ""),
                "current_status": row.get("extraction_review_status", ""),
                "issue_type": ";".join(issues),
                "what_to_check": what_to_check_for_issues(issues),
                "recommended_default_decision": recommended_decision(issues),
                "yossi_decision": "",
                "yossi_notes": "",
            }
        )
    return sorted(items, key=sort_key)


def derive_scope(rows: list[dict[str, str]], provided_scope: str | None) -> str:
    if provided_scope:
        return provided_scope
    refs = [row.get("ref", "") for row in rows if row.get("ref")]
    if not refs:
        return "unknown scope"
    return refs[0] if refs[0] == refs[-1] else f"{refs[0]}-{refs[-1]}"


def write_csv(path: Path, items: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REQUIRED_COLUMNS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(items)


def write_xlsx(path: Path, items: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "XLSX output requires openpyxl, but it is not installed. "
            "Use the Markdown review sheet or UTF-8-BOM CSV, or install openpyxl before requesting --output-xlsx."
        ) from error

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Yossi Review"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.rightToLeft = True

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    hebrew_alignment = Alignment(wrap_text=True, vertical="top", horizontal="right", readingOrder=2)

    sheet.append(list(REQUIRED_COLUMNS))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_top

    for item in items:
        sheet.append([item.get(column, "") for column in REQUIRED_COLUMNS])

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = hebrew_alignment if cell.column_letter == "C" else wrap_top

    column_widths = {
        "A": 26,
        "B": 16,
        "C": 28,
        "D": 34,
        "E": 38,
        "F": 38,
        "G": 18,
        "H": 18,
        "I": 24,
        "J": 28,
        "K": 48,
        "L": 24,
        "M": 22,
        "N": 36,
    }
    for index, column in enumerate(REQUIRED_COLUMNS, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = column_widths.get(letter, max(14, len(column) + 2))

    workbook.save(path)


def markdown_table(items: list[dict[str, str]]) -> str:
    if not items:
        return "No rows selected for this review sheet."
    lines = [
        "| Row ID | Ref | Hebrew phrase | Linear translation | Issue type | What to check | Default | Yossi decision | Notes |",
        "|---|---|---|---|---|---|---|---|---|",
    ]
    for item in items:
        lines.append(
            "| {row_id} | {ref} | {hebrew_phrase} | {linear_translation} | {issue_type} | {what_to_check} | "
            "{recommended_default_decision} |  |  |".format(**item)
        )
    return "\n".join(lines)


def write_markdown(path: Path, items: list[dict[str, str]], rows: list[dict[str, str]], scope: str, map_path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    clean_count = sum(1 for item in items if item["issue_type"] == "clean_sample")
    attention_count = len(items) - clean_count
    allowed_decisions = ", ".join(f"`{decision}`" for decision in ALLOWED_DECISIONS)
    text = f"""# Yossi Source-to-Skill Review Sheet

## Review Summary

- Scope: {scope}
- Source map: `{repo_relative(map_path)}`
- Source map row count: {len(rows)}
- Rows needing review in this sheet: {attention_count}
- Clean sample rows in this sheet: {clean_count}
- Allowed decisions: {allowed_decisions}

Mark each row with one of the allowed decisions. If everything is accurate, use `verified`.

Your job is not to approve questions. Your job is only to confirm whether the source-to-skill extraction is accurate enough to mark this slice extraction-verified. All question, preview, reviewed-bank, runtime, and student-facing gates remain closed.

## What Yossi Is Confirming

- Hebrew phrase text is correct.
- English/source translation is aligned to the correct Hebrew phrase.
- Phrase boundaries and joins are reasonable.
- Parentheticals are attached to the correct phrase.
- Skill/classification is reasonable for source-to-skill planning.
- Any row that should remain source-only is identified.
- Any correction needed before extraction verification is clear.

## What Yossi Is Not Approving

- Not question approval.
- Not protected-preview approval.
- Not reviewed-bank approval.
- Not runtime approval.
- Not student-facing release.
- Not answer-key approval.
- Not generated-question approval.

## Review Rows

{markdown_table(items)}
"""
    path.write_text(text, encoding="utf-8")


def generate_review_sheet(
    map_path: Path,
    output_md: Path,
    output_csv: Path,
    output_xlsx: Path | None = None,
    scope: str | None = None,
    max_clean_samples: int = 3,
) -> dict[str, object]:
    rows = read_tsv(map_path)
    review_scope = derive_scope(rows, scope)
    items = build_review_items(rows, map_path, max_clean_samples)
    write_markdown(output_md, items, rows, review_scope, map_path)
    write_csv(output_csv, items)
    summary: dict[str, object] = {
        "map": repo_relative(map_path),
        "output_md": repo_relative(output_md),
        "output_csv": repo_relative(output_csv),
        "scope": review_scope,
        "source_row_count": len(rows),
        "review_row_count": len(items),
        "clean_sample_count": sum(1 for item in items if item["issue_type"] == "clean_sample"),
    }
    if output_xlsx is not None:
        write_xlsx(output_xlsx, items)
        summary["output_xlsx"] = repo_relative(output_xlsx)
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate concise Yossi Markdown/CSV review sheets.")
    parser.add_argument("--map", required=True)
    parser.add_argument("--output-md", required=True)
    parser.add_argument("--output-csv", required=True)
    parser.add_argument("--output-xlsx", help="Optional XLSX review sheet output. Requires openpyxl if requested.")
    parser.add_argument("--scope")
    parser.add_argument("--max-clean-samples", type=int, default=3)
    args = parser.parse_args()

    try:
        summary = generate_review_sheet(
            ROOT / args.map,
            ROOT / args.output_md,
            ROOT / args.output_csv,
            output_xlsx=ROOT / args.output_xlsx if args.output_xlsx else None,
            scope=args.scope,
            max_clean_samples=args.max_clean_samples,
        )
    except RuntimeError as error:
        parser.exit(status=1, message=f"error: {error}\n")
    print(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
